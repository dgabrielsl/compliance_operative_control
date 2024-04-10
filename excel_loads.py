import os
import sqlite3

from PyQt6.QtWidgets import *

import openpyxl

import re
from datetime import datetime

from catch_cell_data import *

class Excel(QWidget):
    def __init__(self):
        super().__init__()

    def load_sysde(self):
        self.path_a = QFileDialog.getOpenFileName(filter=('*.xlsx'))
        self.path_a = self.path_a[0]

        if self.path_a != '':
            wb = openpyxl.load_workbook(self.path_a)
            ws = wb.worksheets[0]

            ws.delete_rows(1,4)

            char_1 = char_2 = char_3 = ''

            for i in range(ws.max_column):
                i += 1
                if ws.cell(1,i).value == 'Identificación' or ws.cell(1,i).value == 'Identificacion': char_1 = ws.cell(1,i).column_letter
                if ws.cell(1,i).value == 'Email': char_2 = ws.cell(1,i).column_letter
                if ws.cell(1,i).value == 'Teléfono celular' or ws.cell(1,i).value == 'Telefono celular': char_3 = ws.cell(1,i).column_letter

            self.records_from_sysde = []

            for i in range(int(ws.max_row) + 1):
                if i > 1:
                    line = []
                    insert = f'{ws[char_1+str(i)].value}'
                    insert = insert.replace('-','')
                    line.append(str(insert))
                    line.append(f'{ws[char_2+str(i)].value}'.lower())
                    line.append(str(f'{ws[char_3+str(i)].value}'))
                    self.records_from_sysde.append(line)

            QMessageBox.information(self,
                'DeskPyL COM',
                f'\n{len(self.records_from_sysde)} registros nuevos preparados para cargar.\t\t\nIndique una etiqueta para los nuevos registros y selecciones "Guardar".\t\t\n\n📣 La etiqueta se puede repetir.\t\t\n📣 Si el nuevo registro (el número de ID) ya existe, el registro se omitirá.\t\t',
                QMessageBox.StandardButton.Ok,
                QMessageBox.StandardButton.Ok)

    def save_sysde(self):
        tagname_len = len(self.load_sysde_tagname.text())
        if tagname_len > 2 and tagname_len < 99:
            con = sqlite3.connect('hub.db')
            cur = con.cursor()

            timestamp = datetime.now().strftime('%d/%m/%Y %H:%M:%SH')

            try:
                if len(self.records_from_sysde) > 0:

                    for record in self.records_from_sysde:
                        r = f'INSERT INTO sysde VALUES ("{timestamp}", "{self.load_sysde_tagname.text()}", "{record[0]}", "{record[1]}", "{record[2]}")'

                        try:cur.execute(r)
                        except Exception as e: print(e)

                    QMessageBox.information(self, 'DeskPyL', f'\n{len(self.records_from_sysde)} registros nuevos cargados correctamente.\t\t\n', QMessageBox.StandardButton.Ok, QMessageBox.StandardButton.Ok)
                    self.load_sysde_tagname.setText('')
                    self.records_from_sysde.clear()

                else: QMessageBox.information(self, 'DeskPyL', f'\nNo se han precargado datos nuevos para procesar.\t\t\nPor favor cargue datos del reporte de SYSDE para continuar\t\t\n', QMessageBox.StandardButton.Ok, QMessageBox.StandardButton.Ok)

            except Exception as e: QMessageBox.information(self, 'DeskPyL', f'\nNo se han precargado datos nuevos para procesar.\t\t\nPor favor cargue datos del reporte de SYSDE para continuar\t\t\n', QMessageBox.StandardButton.Ok, QMessageBox.StandardButton.Ok)

            con.commit()
            con.close()

        else: QMessageBox.warning(self, 'DeskPyL', '\nEl nombre de la etiqueta debe ser mayor a 3 y menor que 99 caracteres (letras).\t\t\n', QMessageBox.StandardButton.Ok, QMessageBox.StandardButton.Ok)

    def load_hds(self):
        self.path_b = QFileDialog.getOpenFileName(filter=('*.xlsx'))
        self.path_b = self.path_b[0]

        if self.path_b != '':
            wb = openpyxl.load_workbook(self.path_b)
            ws = wb.worksheets[0]

            self.helpdesk = ''
            self.status = ''
            self.fname = ''
            self.author = ''
            self.assigned_to = ''
            self.updated = ''
            self.identification = ''
            self.document = ''
            self.class_case = ''
            self.deadline = ''
            self.product = ''
            self.result = ''
            self.customer_answer = ''
            self.code = ''
            self.income_source = ''
            self.warning_amount = ''
            self.customer_profile = ''
            self.notif_type = ''
            self.contact_type = ''

            for i in range(ws.max_column):
                i += 1
                value = ws.cell(1,i).value.lower()
                value = value.replace(':','').replace('.','')
                value = value.replace('á','a').replace('é','e').replace('í','i').replace('ó','o').replace('ú','u')

                if value.__contains__('#'): self.helpdesk = ws.cell(1,i).column_letter
                elif value.__contains__('estado'): self.status = ws.cell(1,i).column_letter
                elif value.__contains__('asunto'): self.fname = ws.cell(1,i).column_letter
                elif value.__contains__('autor'): self.author = ws.cell(1,i).column_letter
                elif value.__contains__('asignado a'): self.assigned_to = ws.cell(1,i).column_letter
                elif value.__contains__('actualizado'): self.updated = ws.cell(1,i).column_letter
                elif value.__contains__('cedula'): self.identification = ws.cell(1,i).column_letter
                elif value.__contains__('pagare'): self.document = ws.cell(1,i).column_letter
                elif value.__contains__('tipo de caso'): self.class_case = ws.cell(1,i).column_letter
                elif value.__contains__('fecha de prorroga'): self.deadline = ws.cell(1,i).column_letter
                elif value.__contains__('producto'): self.product = ws.cell(1,i).column_letter
                elif value.__contains__('resultado de gestion'): self.result = ws.cell(1,i).column_letter
                elif value.__contains__('respuesta del cliente'): self.customer_answer = ws.cell(1,i).column_letter
                elif value.__contains__('codigo de cliente'): self.code = ws.cell(1,i).column_letter
                elif value.__contains__('origen de fondos'): self.income_source = ws.cell(1,i).column_letter
                elif value.__contains__('monto de la alerta'): self.warning_amount = ws.cell(1,i).column_letter
                elif value.__contains__('perfil del cliente'): self.customer_profile = ws.cell(1,i).column_letter
                elif value.__contains__('tipo de notificacion'): self.notif_type = ws.cell(1,i).column_letter
                elif value.__contains__('tipo de contacto'): self.contact_type = ws.cell(1,i).column_letter

        else: QMessageBox.warning(self, 'DeskPyL', '\nNo se ha cargado ningún archivo.\t\t\n', QMessageBox.StandardButton.Ok, QMessageBox.StandardButton.Ok)

        if self.path_b != '':
            self.datalake = []
            self.insert = ''

            con = sqlite3.connect('hub.db')
            cur = con.cursor()
            cur.execute('SELECT * FROM dictionary')
            res = cur.fetchall()

            self.dict_instructions = []

            for r in res:
                for rr in r:
                    self.dict_instructions.append(rr)

            con.close()

            for i in range(int(ws.max_row) + 1):
                if i > 1:
                    line = []

                    self.insert = f'{ws[self.helpdesk+str(i)].value}'
                    line.append(self.insert)

                    self.insert = f'{ws[self.status+str(i)].value}'
                    line.append(self.insert)

                    self.insert = f'{ws[self.fname+str(i)].value}'
                    Cell.ccd_fname(self)
                    line.append(self.insert)

                    self.insert = f'{ws[self.author+str(i)].value}'
                    Cell.ccd_full_name_titled(self)
                    line.append(self.insert)

                    self.insert = f'{ws[self.assigned_to+str(i)].value}'
                    Cell.ccd_full_name_titled(self)
                    line.append(self.insert)

                    self.insert = f'{ws[self.updated+str(i)].value}'
                    Cell.ccd_updated(self)
                    line.append(self.insert)

                    self.insert = f'{ws[self.identification+str(i)].value}'
                    line.append(self.insert)

                    self.id_match_drop_rule = self.insert

                    self.insert = f'{ws[self.document+str(i)].value}'
                    Cell.ccd_document(self)
                    line.append(self.insert)

                    self.insert = f'{ws[self.class_case+str(i)].value}'
                    line.append(self.insert)

                    self.insert = f'{ws[self.deadline+str(i)].value}'
                    Cell.ccd_deadline(self)
                    line.append(self.insert)

                    self.insert = f'{ws[self.product+str(i)].value}'
                    Cell.ccd_product(self)
                    line.append(self.insert)

                    self.insert = f'{ws[self.result+str(i)].value}'
                    Cell.ccd_result(self)
                    line.append(self.insert)

                    self.insert = f'{ws[self.customer_answer+str(i)].value}'
                    Cell.ccd_customer_answer(self)
                    line.append(self.insert)

                    self.insert = f'{ws[self.code+str(i)].value}'
                    Cell.ccd_code(self)
                    line.append(self.insert)

                    self.insert = f'{ws[self.income_source+str(i)].value}'
                    Cell.ccd_income_source(self)
                    line.append(self.insert)

                    self.insert = f'{ws[self.warning_amount+str(i)].value}'
                    Cell.ccd_warning_amount(self)
                    line.append(self.insert)

                    self.insert = f'{ws[self.customer_profile+str(i)].value}'
                    Cell.ccd_customer_profile(self)
                    line.append(self.insert)

                    self.insert = f'{ws[self.notif_type+str(i)].value}'
                    Cell.ccd_notif_type(self)
                    line.append(self.insert)

                    self.insert = f'{ws[self.contact_type+str(i)].value}'
                    Cell.ccd_contact_type(self)
                    line.append(self.insert)

                    self.datalake.append(line)

            QMessageBox.information(self,
                'DeskPyL COM',
                f'\n{len(self.datalake)} registros nuevos preparados para cargar.\t\t\nIndique una etiqueta para los nuevos registros y selecciones "Guardar".\t\t\n📣 Para los registros de reportes de HDs, la etiqueta debe ser única.',
                QMessageBox.StandardButton.Ok,
                QMessageBox.StandardButton.Ok)

    def save_hdsreport(self):
        tagname_len = len(self.load_hds_tagname.text())
        tagname_str = self.load_hds_tagname.text()

        if tagname_len > 2 and tagname_len < 99:
            con = sqlite3.connect('hub.db')
            cur = con.cursor()
            cur.execute('SELECT * FROM core WHERE tag_name = ?', (tagname_str,))
            res = cur.fetchone()

            if res == None or res == 'None':
                timestamp = datetime.now().strftime('%d/%m/%Y %H:%M:%SH')

                '''
                TABLE hub.db[core]:
                    CREATED                     sys: timestamp
                    TAG_NAME                    sys: tagname_str
                    SYSTEM_ASSIGNED_TO          str: Pendiente
                    HELPDESK                    dl[0]                   &&       HELPDESK                  dl[0]
                    ID                          dl[6]                   &&       STATUS                    dl[1]
                    DOCUMENT                    dl[7]                   &&       FNAME                     dl[2]
                    CODE                        dl[13]                  &&       AUTHOR                    dl[3]
                    CLASS_CASE                  dl[8]                   &&       ASSIGNED_TO               dl[4]
                    STATUS                      dl[1]                   &&       UPDATED                   dl[5]
                    PRODUCT                     dl[10]                  &&       ID                        dl[6]
                    INCOME_SOURCE               dl[14]                  &&       DOCUMENT                  dl[7]
                    WARNING_AMOUNT              dl[15]                  &&       CLASS_CASE                dl[8]
                    CUSTOMER_PROFILE            dl[16]                  &&       DEADLINE                  dl[9]
                    NOTIFICATION_TYPE           dl[17]                  &&       PRODUCT                   dl[10]
                    CONTACT_TYPE                dl[18]                  &&       RESULT                    dl[11]
                    CUSTOMER_ANSWER             dl[12]                  &&       CUSTOMER_ANSWER           dl[12]
                    AUTHOR                      dl[3]                   &&       CODE                      dl[13]
                    ASSIGNED_TO                 dl[4]                   &&       INCOME_SOURCE             dl[14]
                    RESULT                      dl[11]                  &&       WARNING_AMOUNT            dl[15]
                    UPDATED                     dl[5]                   &&       CUSTOMER_PROFILE          dl[16]
                    DEADLINE                    dl[9]                   &&       NOTIFICATION_TYPE         dl[17]
                    FNAME                       dl[2]                   &&       CONTACT_TYPE              dl[18]
                '''

                for dl in self.datalake:
                    try:
                        record = f'INSERT INTO core VALUES ("{timestamp}", "{tagname_str}", "Pendiente", "{dl[0]}", "{dl[6]}", "{dl[7]}", "{dl[13]}", "{dl[8]}", "{dl[1]}", "{dl[10]}", "{dl[14]}", "{dl[15]}", "{dl[16]}", "{dl[17]}", "{dl[18]}", "{dl[12]}", "{dl[3]}", "{dl[4]}", "{dl[11]}", "{dl[5]}", "{dl[9]}", "{dl[2]}")'
                        cur.execute(record)
                        con.commit()
                    except Exception as e: pass

                QMessageBox.information(self,
                    'DeskPyL COM',
                    f'\n{len(self.datalake)} nuevos registros cargados correctamente.\t\t\n\n📣 Si el nuevo registro (el número de ID) ya existe, el registro se omitirá.\t\t',
                    QMessageBox.StandardButton.Ok,
                    QMessageBox.StandardButton.Ok)

                QMessageBox.information(self,
                    'DeskPyL COM',
                    f'\n🐛 Aviso de bug :(\t\t\n\nDespués de cada carga, debe reiniciar la aplicación para poder administrar las solicitudes no asignadas.\t\t',
                    QMessageBox.StandardButton.Ok,
                    QMessageBox.StandardButton.Ok)

                self.load_hds_tagname.setText('')

            else:
                QMessageBox.information(self,
                    'DeskPyL COM',
                    f'\nLa etiqueta ({res}) ya existe.\t\t\nPor favor indique un identificador único para continuar.\t\t',
                    QMessageBox.StandardButton.Ok,
                    QMessageBox.StandardButton.Ok)

            con.close()

        else: QMessageBox.warning(self, 'DeskPyL', '\nEl nombre de la etiqueta debe ser mayor a 3 y menor que 99 caracteres (letras).\t\t\n', QMessageBox.StandardButton.Ok, QMessageBox.StandardButton.Ok)